from typing import Any
import sys
import config
from openpyxl import load_workbook, Workbook
import pandas as pd
import pprint as pp
import datetime
import json
import attribute, cohort, logger, data_use_limitation, cv_list

conf = config.Config.get_config()
logname = conf["logging"]["logfile"]
log = logger.get_logger(logname)
attr_util = attribute.Attribute(log)
cohort_util = cohort.Cohort(log)
dul_util = data_use_limitation.DataUseLimitation(log)
cv_util = cv_list.CVList(log)


def dictify(s):
    """
    Method to build a dictionary recursively from the dataframe based on grouped columns 
    """
    if s.index.nlevels == 1: return s.to_dict(orient="index")
    return {k: dictify(g.droplevel(0)) for k, g in s.groupby(level=0)}


def build_attribute_list(record, attr_fields, multi_value_attributes, attributes, optional_attributes, build_using_db_values=False) -> list:
    """
    # Returns a list of dictionaries such as:
    # [{'attr_name' : 'sex_assigned_at_birth', 'value' : 'female', 'unit' : None, 'source_value' : 'female'}]
    """
    attr_list = []

    for field in attr_fields:
        
        if build_using_db_values:
            db_field = field
            attribute_obj = attr_util.get_attribute({"attr_name" : db_field})

            # print(f"field {field} returned attribute_obj is: {attribute_obj}")

        # only include attributes if they are populated populated
        if field in record and record[field]:
            source_value = record[field]
            value = source_value
            unit = None

            # Special handling for multi-value attribute fields (ones that allow semicolons)
            if field in multi_value_attributes:
                # If multi-value field, split into multiple attributes
                values = split_field(record, field)
                for value in values:
                    if build_using_db_values:
                        # attr_list.append({"attributes_id" : attribute_obj.id, db_key_value[0] : db_key_value[1], 'value' : str(value), 'unit' : unit, 'source_value' : str(source_value)})
                        attr_list.append({"attributes_id" : attribute_obj.id, 'value' : str(value), 'unit' : unit, 'source_value' : str(source_value)})
                    else:
                        attr_list.append({'attr_name' : field, 'value' : value, 'unit' : unit, 'source_value' : source_value})
                continue

            if build_using_db_values:
                # attr_list.append({"attributes_id" : attribute_obj.id, db_key_value[0] : db_key_value[1], 'value' : value, 'unit' : unit, 'source_value' : source_value})
                attr_list.append({"attributes_id" : attribute_obj.id, 'value' : str(value), 'unit' : unit, 'source_value' : str(source_value)})
            else:
                attr_list.append({'attr_name' : field, 'value' : value, 'unit' : unit, 'source_value' : source_value})

    # After handling the attribute fields in the record, process any attributes that were in the attributes tab
    for attr_name in attributes.keys():
        # Retrieve the attribute object that matches the one specified
        attr = attributes[attr_name]
        attribute_obj = attr_util.get_attribute({"attr_name" : attr_name})
        if not attribute_obj:
            raise Exception(f"Could not find the attribute {attr_name} in the database. Cannot process attribute ...")
        source_value = attr['attribute_value']
        unit = attr['unit'] 

        # Special handling for multi-value attribute fields (ones that allow semicolons)
        if attr_name in multi_value_attributes:
            # If multi-value attr_name, split into multiple attributes
            unstripped_result = source_value.strip().split(";")
            values = [s.strip() for s in unstripped_result]
            for value in values:
                if build_using_db_values:
                    attr_list.append({"attributes_id" : attribute_obj.id, 'value' : str(value), 'unit' : unit, 'source_value' : str(source_value)})
                else:
                    attr_list.append({'attr_name' : attr_name, 'value' : value, 'unit' : unit, 'source_value' : source_value})
        else:
            value = source_value
            if build_using_db_values:
                attr_list.append({"attributes_id" : attribute_obj.id, 'value' : str(value), 'unit' : unit, 'source_value' : str(source_value)})
            else:
                attr_list.append({'attr_name' : attr_name, 'value' : value, 'unit' : unit, 'source_value' : source_value})

    # Process any optional attributes that were in the main sheet
    # The optional attributes come in as a key value pair in a dictionary
    for attr_name, source_value in optional_attributes.items():
        # Retrieve the attribute object that matches the one specified
        attribute_obj = attr_util.get_attribute({"attr_name" : attr_name})
        if not attribute_obj:
            raise Exception(f"Could not find the attribute {attr_name} in the database. Cannot process attribute ...")
        unit = None

        # Special handling for multi-value attribute fields (ones that allow semicolons)
        if attr_name in multi_value_attributes:
            # If multi-value attr_name, split into multiple attributes
            unstripped_result = source_value.strip().split(";")
            values = [s.strip() for s in unstripped_result]
            for value in values:
                if build_using_db_values:
                    attr_list.append({"attributes_id" : attribute_obj.id, 'value' : str(value), 'unit' : unit, 'source_value' : str(source_value)})
                else:
                    attr_list.append({'attr_name' : attr_name, 'value' : value, 'unit' : unit, 'source_value' : source_value})
        else:
            value = source_value
            if build_using_db_values:
                attr_list.append({"attributes_id" : attribute_obj.id, 'value' : str(value), 'unit' : unit, 'source_value' : str(source_value)})
            else:
                attr_list.append({'attr_name' : attr_name, 'value' : value, 'unit' : unit, 'source_value' : source_value})

    log.debug(f"Returing attributes list: {attr_list}")
    return attr_list

def split_field(record, field_name):

    result = []
    if field_name in record and record[field_name]:
        value = str(record[field_name])
        unstripped_result = value.strip().split(";")
        # print(f"unstripped_result: {unstripped_result}")
        result = [s.strip() for s in unstripped_result]
    # print(f"Returning {result}")
    return result

def get_excel_worksheet_name_case_insensitive(wb, sheet_name):
    # get the worksheet names, regardless of case
    for sheetname in wb.sheetnames:
        if sheetname.lower() == sheet_name.lower():
            return sheetname
    return None
        
class AssetsLoader:

    def __init__(self, log):   
        self.log = log
        attr_util = attribute.Attribute(log)

    def convert_key_value_to_attr_with_unit(self, key, value, unit = None):
        attr = {}
        attr['key'] = key
        attr['value'] = value
        attr['unit'] = unit
 
        # Look up to see if the key has a cv associated with it, in which case set it
        cv_entry = cv_util.get_cv({'term_name' : key}) 
        if cv_entry:
            attr["cv_list_id"] = cv_entry.id
        else:
            attr['cv_list_id'] = None
       
        return attr

    def convert_key_value_to_attr(self, key, value):
        attr = {}
        attr['key'] = key
        attr['value'] = value
 
        # Look up to see if the key has a cv associated with it, in which case set it
        cv_entry = cv_util.get_cv({'term_name' : key}) 
        if cv_entry:
            attr["cv_list_id"] = cv_entry.id
        else:
            attr['cv_list_id'] = None
       
        return attr

    def generate_sample_dict(self, record, num_subject_names, subject_objs, is_human, proj_obj, anatomies, parent_sample_map, 
                             sample_attribute_fields, sample_attributes, optional_attributes, multi_value_attributes, for_db_comparison=False):
        """
        Generate the sample dictionary for the parameters passed to this method. The method will also build
        the attributes list from the record as needed.
        """
        smp = {}

        if record.get("alt_sample_id"):
            smp["alt_id"] = record["alt_sample_id"] 
        
        smp["project_id"] = proj_obj.id
        smp["sample_name"] = record["sample_name"]
        smp["sample_type"] = record["sample_type"]
        smp["attributes"] = build_attribute_list(record, sample_attribute_fields, multi_value_attributes, sample_attributes, optional_attributes, build_using_db_values=for_db_comparison)

        if record.get("sample_source_id"):
            smp["source_sample_id"] = record["sample_source_id"]

        if record.get("sample_source"):
            smp["sample_source"] = record["sample_source"]

        if anatomies:
            if for_db_comparison:
                smp["anatomies"] = [anatomy["name"] for anatomy in anatomies]
            else:
                smp["anatomies"] = [anatomy["id"] for anatomy in anatomies]
       
        if record.get("parent_entity_id"):
            smp["sample_assoc_sample_parent"] = [{"parent_sample_id" : parent_sample_obj.id, "relationship" : "parent", "root_sample_id" : None}
                                                     for parent_sample_obj in parent_sample_map.values()]
        
        # TODO: Check to see if this is still valid
        if num_subject_names == 1:
            smp["sbj_ids"] = [subject_objs[0].id]
        elif num_subject_names > 1:
            smp["sbj_ids"] = [subject_obj.id for subject_obj in subject_objs]
        
        return smp

    def generate_library_dict(self, record, library_id, parent_sample_id, parent_library_id, modality_id, assay_id, specimen_type_id, technique_id, project_id, 
                              library_name, is_aliquot, library_attribute_fields, library_attributes, optional_attributes, library_pool_id, multi_value_attributes, for_db_comparison=False):
        """
        Generate the library dictionary for the parameters passed to this method. The method will also build
        the attributes list from the record as needed.
        """

        library = {}
        library["lib_id"] = library_id
        library["modality_id"] = modality_id
        library["assay_id"] = assay_id
        library["specimen_type_id"] = specimen_type_id
        library["technique_id"] = technique_id
        library["project_id"] = project_id
        library["library_name"] = library_name
        library["library_type"] = "aliquot" if is_aliquot else "library"
        if record.get("alt_library_id"):
            library["alt_id"] = record["alt_library_id"] 

        library["lib_assoc_lib_pool"] = library_pool_id

        # As this library has a parent library set the association so it will create the associations
        if parent_library_id:
            log.info(f"Parent is library, so creating parent child relationship")
            library["lib_assoc_lib_parent"] = parent_library_id
        else:
            # As the parent is a sample, set the sample_id
            library["sample_id"] = parent_sample_id


        library["attributes"] = build_attribute_list(record, library_attribute_fields, multi_value_attributes, library_attributes, optional_attributes, build_using_db_values=for_db_comparison)
        
        return library

    def generate_file_dict(self, data_type_obj, file_type_obj, analysis_obj, submission_id, record, proj_obj, library_obj, 
                           parent_file_ids, collection_short_name, file_attrs, dul_obj, version, for_db_comparison=False):
        """
        This function takes the parameters needed to build a dictionary for a file object that can be used by the 
        db util to add/update a record
        """
        file = {}
        file['data_type_id'] = data_type_obj.id
        file['file_format_id'] = file_type_obj.id
        file['submission_id'] = submission_id
        file['file_name'] = record['file_name']
        file['md5'] = record['md5_checksum']
        file['project_id'] = proj_obj.id
        if record.get('size'):
            file['size'] = record['size']
        if record.get('mtime'):
            file['mtime'] = record['mtime']
        file['version'] = version

        # if analysis_obj, add in the analysis ID and the association to the 
        # analysis attributes
        if analysis_obj:
            file['analysis_id'] = analysis_obj.id
            file['analysis_attributes'] = [
                {
                    "name" : "genome_build",
                    "value" : record["genome_build"],
                    "analysis_id" : analysis_obj.id
                },
                {
                    "name" : "gene_set_release",
                    "value" : record["gene_set_release"],
                    "analysis_id" : analysis_obj.id
                }
            ]

        # if file attributes are specified, set up the file_attribute
        if file_attrs:
            file['file_attributes'] = file_attrs

        # If there are any parent file ids in the manifest
        # TODO: Future (round 2?) could add validation to make sure these two files should be connected and wasn't an excel drag or copy/paste issue.
        if parent_file_ids:
            parent_associations = []
            for parent_file_id in parent_file_ids:
                parent_associations.append({
                    "parent_file_id" : parent_file_id,
                    "relationship" : "derived"
                })
            file["file_parents"] = parent_associations

        # Connect the lib and aliquot to the file
        if library_obj:
            # Connect any libs to the file; takes library_name instead of lib.id because in order to use this file
            # dict in a compare_to() call, we need to use the same values as those returned from Assets when retrieving
            # a file with the 'libraries' association
            if for_db_comparison:
                file["libraries"] = [library_obj.library_name]
            else: # for (for db insert, need the IDs)
                file["libraries"] = [library_obj.id]

        # TODO: Should I set these?
        # file['alt_id'] = ?
        # file['sha256'] = ?

        file["data_use_limitations"] = [dul_obj.id]
    
        file['collections'] = [collection_short_name]
        self.log.debug(f"File dictionary: {file}")
        return file

    def worksheet_to_clean_list_of_attrs(self, worksheet, first_expected_fields_in_sheets):
        """ Convert the openpyxl worksheet to a list of dicts organized by filename.
            While at it, strip any whitespaces from the cell values.
        """
        primary_key = None
        attribute = None
        data = {}
        attrs = None
        attr_fields = None
        for row_num, row in enumerate(worksheet.iter_rows(values_only=True)):
            self.log.debug(f"Processing row: {row}")
            if row_num == 0:
                # extra validation check:
                if row and row[0]:
                    value = row[0]
                    if value.lower() not in first_expected_fields_in_sheets:
                        raise ValueError("Expected the header row to appear as the first row of the input file.")
                    else:
                        header_row = row
                else:
                    raise ValueError("Expected the header row to appear as the first row of the input file.")
                continue
            clean_values = [cell.strip() if isinstance(cell, str) else cell for cell in row]

            # If the key value has changed then start a new dictionary
            # Just in case the sheets is not sorted check to see if a dict already exists before
            # creating new one
            if clean_values[0] != primary_key:
                self.log.debug(f"Primary key is changing from {primary_key} to {clean_values[0]}")
                primary_key = clean_values[0]

                # Check to see if a dictionary for this key already exists
                if primary_key in data:
                    attrs = data[primary_key]
                else:
                    attrs = {}
                    data[primary_key] = attrs

            # Add the fields in this row as the attribute values for the primary key
            attribute = row[1]
            attr_fields = {}
            attr_fields['value'] = row[2]
            attr_fields['unit'] = row[3]
            attr_fields['cv_term'] = row[4]
            attrs[attribute] = attr_fields

        self.log.debug(f"Attributes: {data}")
        return data

    def create_dict_of_attrs(self, df, key_columns):
        """
        This method takes a data frame that has been uniquefied across the key columns
        and generates a dict of dictionaries
        """
        df = df.set_index(key_columns)
        grouped_data = dictify(df) 

        self.log.debug(f"Dictionary with grouped data: {grouped_data}")
        return grouped_data

    def parse_subject_attributes(self, workbook_path, sheet_name):
        """
        Parse the subject attributes tab and only returns attributes which are not
        related to events.
        """
        self.log.debug(f"Processing sheet: {sheet_name}")

        excel_file = pd.ExcelFile(workbook_path)
        if not sheet_name in excel_file.sheet_names:
            self.log.warn(f"Sheet with name: {sheet_name} does not exist. Returning empty dict ...")
            return {}

        df = pd.read_excel(excel_file, sheet_name) 

        # As we are only looking at subject specific attributes filter out any that
        # are associated with events
        # print(f"Data frame before filtering out empty events: {df}")
        df = df.fillna('')
        non_event_cols = ["subject_name", "attribute_name", "attribute_value", "unit"]
        df = df[(df['event_name'].isnull()) | (df['event_name'].str.len() == 0)][non_event_cols]
        # print(f"Data frame after filtering out empty events: {df}")
        key_columns = ["subject_name", "attribute_name"]
        grouped_data = self.create_dict_of_attrs(df, key_columns)

        return grouped_data

    def parse_subject_event_attributes(self, workbook_path, sheet_name):
        """
        Parse the subject attributes tab and only returns attributes which are
        related to events.
        """
        self.log.debug(f"Processing sheet: {sheet_name}")

        excel_file = pd.ExcelFile(workbook_path)
        if not sheet_name in excel_file.sheet_names:
            self.log.warn(f"Sheet with name: {sheet_name} does not exist. Returning empty dict ...")
            return {}

        df = pd.read_excel(excel_file, sheet_name)

        # As we are only looking at subject specific attributes filter out any that
        # are associated with events
        # print(f"Data frame before filtering out non-empty events: {df}")
        df = df.fillna('')
        non_event_cols = ["subject_name", "event_name", "attribute_name", "attribute_value", "unit"]
        df = df[~(df['event_name'].isnull()) | (df['event_name'].str.len() == 0)][non_event_cols]
        # print(f"Data frame after filtering out empty events: {df}")
        key_columns = ["subject_name", "event_name","attribute_name"]
        grouped_data = self.create_dict_of_attrs(df, key_columns)

        return grouped_data


    def worksheet_to_clean_dict_of_attrs(self, workbook_path, sheet_name, key_columns):
        """ Convert the openpyxl worksheet to a dictionary of dictionaries organized by the specified keys.
            While at it, strip any whitespaces from the cell values.
        """
        log.debug(f"Processing sheet: {sheet_name}")
        excel_file = pd.ExcelFile(workbook_path)
        if not sheet_name in excel_file.sheet_names:
            self.log.debug(f"Sheet with name: {sheet_name} does not exist. Returning empty dict ...")
            return {}

        df = pd.read_excel(excel_file, sheet_name, na_values=["NA"]) 
        df = df.fillna('')

        grouped_data = {}
        # Group by the specified columns and convert each group to a dictionary
        # grouped_data = df.groupby(key_columns).apply(lambda x: x.set_index(key_columns).to_dict(orient='index')).to_dict(orient='index')
        # pp.pprint(grouped_data)
        grouped_data = self.create_dict_of_attrs(df, key_columns)

        self.log.debug(f"Data: {df} setting index on columns: {key_columns}")
        return grouped_data

    def worksheet_to_clean_list_of_dicts(self, worksheet, first_expected_fields_in_sheets) -> 'list[dict]':
        """ Convert the openpyxl worksheet to a list of dicts.
            While at it, strip any whitespaces from the cell values.
        """
        header_row = None
        data = []
        for row_num, row in enumerate(worksheet.iter_rows(values_only=True)):
            if row_num == 0:
                # extra validation check:
                if row and row[0]:
                    value = row[0]
                    if value.lower() not in first_expected_fields_in_sheets:
                        raise ValueError("Expected the header row to appear as the first row of the input file.")
                    else:
                        header_row = row
                else:
                    raise ValueError("Missing first row.")
                continue

            # Excel is sometimes storing values as datetime.time or doing other automatic detection and conversion of cells
            # This is problematic in some cases. We usually just want everyting to be a string
            # Added conversion of datetime.time objects due to an issue when trying to insert values to the DB
            
            # trim strings and convert ints to strings
            cleaned_string_values = [cell.strip() if isinstance(cell, str) else str(cell) if isinstance(cell, int) else cell for cell in row]
            # cleaned_string_values = [cell.strip() if isinstance(cell, str) else cell for cell in row]

            # trim strings; convert various datatypes (numeric, boolean, but not datetime.time) to strings
            # cleaned_string_values = [cell.strip() if isinstance(cell, str) else str(cell) if cell is not None and not isinstance(cell, datetime.time) else cell for cell in row]

            # convert datetime.time instances to strings
            clean_str_and_date_values = [cell.strftime("%H:%M:%S") if isinstance(cell, datetime.time) else cell for cell in cleaned_string_values]

            clean_values = dict(zip(header_row, clean_str_and_date_values))
            data.append(clean_values)
        return data

    def is_json(self, s):
        import ast
        try:
            json.loads(ast.literal_eval(s))
        except (ValueError, SyntaxError) as e:
            print(e)
            return False
        return True

    def print_report(self, report, output_file):
        
        import csv
        keys = report[1][3].keys()
        with open(output_file, 'a', newline='') as csvfile:
            # writer = csv.writer(csvfile, delimiter=',')
                                # quotechar='|', quoting=csv.QUOTE_MINIMAL)
            writer = csv.writer(csvfile)
            for row_num, record in enumerate(report):
                # if is_json(record[3]):
                #     print("IS JSON")
                # else:
                #     print("NOT JSON")
                orig_values = record[3]
                # if isinstance(record[3], dict):
                #     print("TYPE DICT")
                if row_num == 0:
                    values = list(keys)
                else:
                    values = list(orig_values.values())
                # else:
                #     print("NOT TYPE DICT")
                #     values = [orig_values]

                writer.writerow(record[0:3] + values)

        # with open(output_file, 'w') as f:
        #     for record in report:
        #         print(record, file=f)
        
    def write_report_to_worksheet(self, report, wksheet):
        row_num = 0
        for line in report:
            for i, value in enumerate(line):
                cellref = wksheet.cell(column=i+1, row=row_num+1) # cols and rows start at 1 in Excel
                cellref.value = str(value)
            row_num += 1

    def errors_in_report(self, report):
        for entry in report:
            self.log.debug(f"Error entry: {entry}")
            if entry[1] == False:
                return True
        return False

    def errors_in_reports(self, subject_report, sample_report, library_report, file_report):
        subject_errors = self.errors_in_report(subject_report) 
        print(f"Subject errors: {subject_errors}")
        sample_errors = self.errors_in_report(sample_report) 
        print(f"Sample errors: {sample_errors}")
        library_errors = self.errors_in_report(library_report) 
        print(f"Library errors: {library_errors}")
        file_errors = self.errors_in_report(file_report) 
        print(f"File errors: {file_errors}")
        return subject_errors or sample_errors or file_errors
        # return errors_in_report(subject_report) or errors_in_report(sample_report) or errors_in_report(file_report)
    
    def retrieve_most_restrictive_cohort_and_dul(self, project_name, is_human):
        """
        This method retrieve all the cohorts associated with this project and identifies the most restrictive 
        cohort and returns that cohort object
        """
        log.debug(f"Retrieving the cohorts for project {project_name} that are for human subjects: {is_human}")
        restrictive_cohort = None
        restrictive_dul = None
        cohorts = cohort_util.get_all_cohorts_for_project(project_name)
        log.debug(f"Retrieved cohorts are: {cohorts}")
        
        access_level = 0
        
        for cohort in cohorts:
            log.debug(f"Examining cohort with asset ID: {cohort.coh_id} and cohort name: {cohort.cohort_name} is human: {cohort.is_human}")
            # If the is_human flag of the cohort being examined does not match the parameter then you can skip examining 
            if bool(cohort.is_human) != is_human:
                log.debug(f"The cohort human flag is {cohort.is_human} and does not match the project is_human flag {is_human} so skipping this cohort.")
                continue

            # Retrieve the DUL associated for non-summary files for this cohort
            duls = dul_util.get_all_duls_for_cohort_summary_flag(cohort.id, summary_flag=False)
            log.debug(f"Retrieved the following DULs: {duls}")

            for dul in duls: 
                log.debug(f"Processing the DUL: {dul} current access level {access_level}.")

                if dul.access == "restricted_embargo":
                    # This is the most retrictive cohort so use it
                    access_level = 4
                    restrictive_cohort = cohort
                    restrictive_dul = dul
                    break
                elif dul.access == "restricted":
                    if access_level > 3:
                        continue
                    else:
                        access_level = 3
                        restrictive_cohort = cohort
                        restrictive_dul = dul
                elif dul.access == "open_embargo":
                    if access_level > 2:
                        continue
                    else:
                        access_level = 2
                        restrictive_cohort = cohort
                        restrictive_dul = dul
                else:
                    if access_level > 1:
                        continue
                    else:
                        access_level = 1
                        restrictive_cohort = cohort
                        restrictive_dul = dul

        log.debug(f"Returning the most restrictive cohort: {restrictive_cohort} and DUL {restrictive_dul}")

        return restrictive_cohort, restrictive_dul

    def is_library_human(self, library_name):
        """
        A method that is a dummy impmentation now, but one that will call the function to 
        retrieve all the taxonomies associated with this library to decide is this library
        is from a human sample
        """

        # TODO: Call the recursive function to retreive the is_human status

        return True
